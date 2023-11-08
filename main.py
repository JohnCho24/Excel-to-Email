# Import modules to connect excel to code
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# Import modules to connect code to mail server
from email.mime.text import MIMEText 
from email.mime.image import MIMEImage 
from email.mime.application import MIMEApplication 
from email.mime.multipart import MIMEMultipart 
import smtplib 
import os 

# load excel file
wb = load_workbook('hello.xlsx')

# Assign sheets to vars
ws1 = wb['Schedule']
ws2 = wb['Master List']

# Get value of day by input
date = input('What is the day you are searching for: ')

# Get column name of our specific date
for cell in ws1['2']:
    if cell.value == date:
        col = get_column_letter(cell.column)
        break

# Create lists for names and emails respectively
names = []
emails = []

# For each name in the schedule, place in names and emails so that it correlates
for cell in ws1[str(col)]:
    name = cell.value
    # Ignore if font colour is red
    if cell.font.color.rgb != 'FFFF0000':
        for cell2 in ws2['J']:
            if name == cell2.value:
                names.append(name)
                emails.append(ws2[f'M{cell2.row}'].value)
                break

# compile all emails into a list

# initialize connection to email server, here we use hotmail
smtp = smtplib.SMTP('smtp.office365.com', 587) 
smtp.ehlo() 
smtp.starttls() 

# Prompt the sender email and password
sender_email = input('Enter your email: ')
sender_pass = input('Enter your password: ')

# Login with your email and password 
smtp.login(sender_email, sender_pass) 

# Builds the email
def message(subject="", text="", img=None, attachment=None): 
	
	# build message contents 
	msg = MIMEMultipart() 
	
	# Add Subject 
	msg['Subject'] = subject 
	
	# Add text contents 
	msg.attach(MIMEText(text)) 

	# Check if there exists an image input
	if img is not None: 
		
		# Check whether we have the lists of images or not! 
		if type(img) is not list: 
			
			# if it isn't a list, make it one 
			img = [img] 

		# Now iterate through our list 
		for one_img in img: 
			
			# read the image binary data 
			img_data = open(one_img, 'rb').read() 
			# Attach the image data to MIMEMultipart 
			# using MIMEImage, we add the given filename use os.basename 
			msg.attach(MIMEImage(img_data, 
								name=os.path.basename(one_img))) 

	# Check if there exists an attachment input
	if attachment is not None: 
		
		# Check whether we have the lists of attachments or not! 
		if type(attachment) is not list: 
			
			# if it isn't a list, make it one 
			attachment = [attachment] 

		for one_attachment in attachment: 

			with open(one_attachment, 'rb') as f: 
				
				# Read in the attachment 
				# using MIMEApplication 
				file = MIMEApplication( 
					f.read(), 
					name=os.path.basename(one_attachment) 
				) 
			file['Content-Disposition'] = f'attachment; filename="{os.path.basename(one_attachment)}"' 
			
			# Add the attachment to our message object 
			msg.attach(file) 
	return msg 

# Take message format
email_subject = "Zoom Happy Hour Reminder"
email_message = "Hey everyone,\n\n\n\nWe are so looking forward to seeing you all at Happy Hour on " + date + "!\n\n\n\nHere is the Zoom link: https://us02web.zoom.us/j/7804952149\n\nMeeting ID: 780 495 2149\n\nTime: 4:00-5:45 PM ET\n\n\n\nPS – If you wouldn’t mind sending me a quick confirmation that you received this, that would be great!  😊\n\n\n\n Thanks,\nValentina"

# Provide inputs for message func
msg = message(email_subject, email_message) 

# List of emails that are being sent to
to = emails

print("Sending...")

# Send mail info
smtp.sendmail(from_addr=sender_email, 
			to_addrs=to, msg=msg.as_string()) 

print("")
# Closing connection 
smtp.quit()
